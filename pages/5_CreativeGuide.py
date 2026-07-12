import streamlit as st
import io
import html
from utils import db
from utils.auth import get_current_user
from utils.ui import set_current_page
from st_click_detector import click_detector
import uuid

set_current_page("creative_guide")

user = get_current_user()
user_email = user.get("email", "") if user else ""
BUCKET = "creative-guides"

guides = db.get_creative_guides()
all_media = db.get_all_media()

media_cat_map = {}
for m in all_media:
    cat = m.get("categories") or {}
    media_cat_map[m["name"]] = cat.get("major_category", "")

guide_map = {}
for g in guides:
    guide_map.setdefault(g["media_name"], {})[g["product_name"]] = g

tab_dl, tab_up = st.tabs(["제작가이드 다운로드", "업로드"])

# ============================
# 다운로드 탭
# ============================
with tab_dl:
    majors = db.get_major_categories()
    col_f1, col_f2, col_or, col_f3 = st.columns([2, 2, 0.4, 3])
    with col_f1:
        major_filter = st.selectbox(
            "대분류", [""] + majors, label_visibility="collapsed",
            key="cg_major", format_func=lambda x: "대분류 선택" if x == "" else x,
        )
    with col_f2:
        if major_filter:
            subs = db.get_sub_categories(major_filter)
            sub_filter = st.selectbox(
                "중분류", [""] + subs, label_visibility="collapsed",
                key="cg_sub", format_func=lambda x: "중분류" if x == "" else x,
            )
        else:
            sub_filter = ""
            st.selectbox("중분류", ["중분류"], label_visibility="collapsed",
                         key="cg_sub_empty", disabled=True)
    with col_or:
        st.markdown("<div style='text-align:center;color:var(--text-muted);font-size:12px;padding-top:8px;'>또는</div>",
                    unsafe_allow_html=True)
    with col_f3:
        search_kw = st.text_input("검색", placeholder="🔍 매체명 직접 검색",
                                   label_visibility="collapsed", key="cg_search")

    any_filter = bool(major_filter or search_kw)
    all_media_names = sorted(set([m["name"] for m in all_media] + list(guide_map.keys())))

    def passes_filter(name: str) -> bool:
        if search_kw and search_kw.lower() not in name.lower():
            return False
        if major_filter and media_cat_map.get(name, "") != major_filter:
            return False
        return True

    filtered_names = [n for n in all_media_names if passes_filter(n)] if any_filter else []

    if "cg_selected" not in st.session_state:
        st.session_state["cg_selected"] = {}

    if not any_filter:
        st.markdown(
            "<div style='margin:40px auto;max-width:480px;background:rgba(0,0,0,0.04);"
            "border-radius:12px;padding:24px 28px;opacity:0.6;text-align:center;'>"
            "<div style='font-size:14px;font-weight:600;margin-bottom:12px;'>Quick Guide</div>"
            "<div style='font-size:13px;color:var(--text-secondary);text-align:center;line-height:2;'>"
            "① 희망하는 매체를 카테고리에서 직접 선택하거나 검색<br>"
            "② 해당 상품 체크<br>"
            "③ 체크 완료된 파일 확인 후 다운로드 버튼 클릭!"
            "</div></div>",
            unsafe_allow_html=True,
        )
    else:
        # 전체 매체 행을 하나의 HTML로 통합
        all_rows_html = []
        for media_name in filtered_names:
            products = guide_map.get(media_name, {})
            if not products:
                continue

            btn_parts = []
            for product_name, guide in sorted(products.items()):
                has_file = bool(guide.get("storage_path"))
                key = (media_name, product_name)
                is_on = st.session_state["cg_selected"].get(key, False)
                pid = html.escape(f"{media_name}||{product_name}", quote=True)
                pn_disp = html.escape(product_name)

                if not has_file:
                    btn_parts.append(
                        f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                        f"box-shadow:0 0 0 0.5px #666 inset;"
                        f"color:#999;opacity:0.45;cursor:not-allowed;"
                        f"display:inline-block;'>{pn_disp}</span>"
                    )
                elif is_on:
                    btn_parts.append(
                        f"<a href='#' id='{pid}' style='text-decoration:none;'>"
                        f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                        f"background:#111;color:#fff;"
                        f"cursor:pointer;display:inline-block;'>✓ {pn_disp}</span></a>"
                    )
                else:
                    btn_parts.append(
                        f"<a href='#' id='{pid}' style='text-decoration:none;'>"
                        f"<span style='padding:6px 14px;font-size:13px;border-radius:8px;"
                        f"box-shadow:0 0 0 0.5px #111 inset;"
                        f"color:#111;cursor:pointer;display:inline-block;'>{pn_disp}</span></a>"
                    )

            row_html = (
                "<div style='display:flex;align-items:center;gap:0;padding:10px 0;"
                "border-bottom:0.5px solid #e0e0e0;'>"
                f"<div style='flex:0 0 80px;font-size:13px;font-weight:600;'>{html.escape(media_name)}</div>"
                "<div style='width:1px;height:32px;background:#ccc;margin:0 12px;flex-shrink:0;'></div>"
                "<div style='display:flex;flex-wrap:wrap;gap:8px;flex:1;'>"
                + "".join(btn_parts) +
                "</div></div>"
            )
            all_rows_html.append(row_html)

        if all_rows_html:
            full_html = "<div>" + "".join(all_rows_html) + "</div>"
            clicked = click_detector(full_html, key="cg_det_all")
            last_key = "_cg_last_all"
            if clicked and "||" in clicked and clicked != st.session_state.get(last_key):
                st.session_state[last_key] = clicked
                decoded = html.unescape(clicked)
                mn, pn = decoded.split("||", 1)
                ck = (mn, pn)
                g = guide_map.get(mn, {}).get(pn)
                if g and bool(g.get("storage_path")):
                    st.session_state["cg_selected"][ck] = not st.session_state["cg_selected"].get(ck, False)
                    st.rerun()

        # 선택 태그 + 다운로드 버튼
        selected = {k: v for k, v in st.session_state["cg_selected"].items() if v}
        if selected:
            tag_html = "".join(
                f"<span style='font-size:12px;padding:4px 12px;border-radius:20px;"
                f"box-shadow:0 0 0 0.5px #999 inset;"
                f"display:inline-block;margin:3px;'>"
                f"{html.escape(mn)} · {html.escape(pn)}</span>"
                for (mn, pn) in selected
            )
            st.markdown(
                f"<div style='margin-top:16px;margin-bottom:8px;'>{tag_html}</div>",
                unsafe_allow_html=True,
            )

            @st.fragment
            def download_fragment():
                sel = {k: v for k, v in st.session_state["cg_selected"].items() if v}
                if not sel:
                    return
                try:
                    import openpyxl
                    from copy import copy
                    merged_wb = openpyxl.Workbook()
                    merged_wb.remove(merged_wb.active)
                    for (mn, pn) in sel:
                        guide = guide_map.get(mn, {}).get(pn)
                        if not guide:
                            continue
                        file_bytes = db.download_from_storage(BUCKET, guide["storage_path"])
                        src_wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                        for sheet_name in src_wb.sheetnames:
                            src_ws = src_wb[sheet_name]
                            new_ws = merged_wb.create_sheet(title=sheet_name[:31])
                            for row in src_ws.iter_rows():
                                for cell in row:
                                    new_ws[cell.coordinate].value = cell.value
                                    if cell.has_style:
                                        new_ws[cell.coordinate].font = copy(cell.font)
                                        new_ws[cell.coordinate].fill = copy(cell.fill)
                                        new_ws[cell.coordinate].border = copy(cell.border)
                                        new_ws[cell.coordinate].alignment = copy(cell.alignment)
                                        new_ws[cell.coordinate].number_format = cell.number_format
                            for col_dim in src_ws.column_dimensions.values():
                                new_ws.column_dimensions[col_dim.index].width = col_dim.width
                            for row_dim in src_ws.row_dimensions.values():
                                new_ws.row_dimensions[row_dim.index].height = row_dim.height
                    buf = io.BytesIO()
                    merged_wb.save(buf)
                    buf.seek(0)
                    st.download_button(
                        "선택한 제작가이드 통합 다운로드",
                        data=buf,
                        file_name="통합_제작가이드.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"다운로드 중 오류: {e}")

            download_fragment()

# ============================
# 업로드 탭
# ============================
with tab_up:
    media_options = sorted(set([m["name"] for m in all_media]))

    def log_action(action, media_name, product_name, old_product_name=None):
        try:
            db.get_client().table("creative_guide_logs").insert({
                "action": action,
                "media_name": media_name,
                "product_name": product_name,
                "old_product_name": old_product_name,
                "user_email": user_email,
            }).execute()
        except Exception as e:
            st.warning(f"로그 기록 실패: {e}")

    col_m, col_act = st.columns([2, 2])
    with col_m:
        up_media = st.selectbox("매체 선택", media_options, key="cg_up_media")
    with col_act:
        up_action = st.selectbox("작업 구분", ["등록", "수정", "삭제"], key="cg_up_action")

    existing_products = sorted(guide_map.get(up_media, {}).keys())

    # ---------- 등록 ----------
    if up_action == "등록":
        up_product = st.text_input("상품명", key="cg_up_product")
        up_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"], key="cg_up_file")

        if st.button("저장", type="primary", key="cg_up_save"):
            if not up_media or not up_product:
                st.error("매체와 상품명을 입력해주세요.")
            else:
                existing = guide_map.get(up_media, {}).get(up_product)
                if up_file:
                    file_bytes = up_file.read()
                    storage_path = f"{uuid.uuid4()}.xlsx"
                    if existing:
                        if existing.get("storage_path"):
                            db.delete_from_storage(BUCKET, existing["storage_path"])
                        db.upload_to_storage(BUCKET, storage_path, file_bytes)
                        db.update_creative_guide(existing["id"], storage_path)
                    else:
                        db.upload_to_storage(BUCKET, storage_path, file_bytes)
                        cat = media_cat_map.get(up_media, "")
                        db.create_creative_guide(up_media, cat, up_product, storage_path)
                else:
                    if not existing:
                        cat = media_cat_map.get(up_media, "")
                        db.create_creative_guide(up_media, cat, up_product, "")
                log_action("등록", up_media, up_product)
                st.session_state["_cg_upload_success"] = f"'{up_media} · {up_product}' 저장 완료."
                st.rerun()

    # ---------- 수정 ----------
    elif up_action == "수정":
        if not existing_products:
            st.info("해당 매체에 등록된 상품이 없습니다.")
        else:
            col_old, col_new = st.columns([2, 2])
            with col_old:
                old_product = st.selectbox("기존 상품명", existing_products, key="cg_up_old")
            with col_new:
                new_product = st.text_input("변경할 상품명", key="cg_up_new")

            if st.button("수정 저장", type="primary", key="cg_up_edit"):
                if not new_product:
                    st.error("변경할 상품명을 입력해주세요.")
                elif new_product in existing_products:
                    st.error("이미 존재하는 상품명입니다.")
                else:
                    guide = guide_map.get(up_media, {}).get(old_product)
                    if guide:
                        db.update_creative_guide_name(guide["id"], new_product)
                        log_action("수정", up_media, new_product, old_product_name=old_product)
                        st.session_state["_cg_upload_success"] = f"'{old_product}' → '{new_product}' 수정 완료."
                        st.rerun()

    # ---------- 삭제 ----------
    elif up_action == "삭제":
        if not existing_products:
            st.info("해당 매체에 등록된 상품이 없습니다.")
        else:
            del_product = st.selectbox("삭제할 상품명", existing_products, key="cg_up_del")
            confirm = st.checkbox("정말 삭제하시겠습니까?", key="cg_up_del_confirm")

            if st.button("삭제", type="primary", key="cg_up_del_btn"):
                if not confirm:
                    st.error("삭제 확인 체크박스를 선택해주세요.")
                else:
                    guide = guide_map.get(up_media, {}).get(del_product)
                    if guide:
                        if guide.get("storage_path"):
                            db.delete_from_storage(BUCKET, guide["storage_path"])
                        db.delete_creative_guide(guide["id"])
                        log_action("삭제", up_media, del_product)
                        st.session_state["_cg_upload_success"] = f"'{up_media} · {del_product}' 삭제 완료."
                        st.rerun()

    msg = st.session_state.pop("_cg_upload_success", None)
    if msg:
        st.success(msg)
